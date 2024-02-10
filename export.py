
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def export(reviews, aboutReviews, TFIDF, placeID, extraInformation):
    book = load_workbook("project.xlsx")
    sheet = book[placeID] if placeID in book.sheetnames else book.create_sheet(
        placeID)

    sheet = book.active
    print("REVIEWS", reviews)
    print("ABOUT", aboutReviews)
    print("VECTOR", TFIDF)

    sheet['C5'] = reviews[0]
    sheet.merge_cells('C5:F8')
    sheet['C9'] = reviews[1]
    sheet.merge_cells('C9:F12')
    sheet['C13'] = reviews[2]
    sheet.merge_cells('C13:F16')
    sheet['C17'] = reviews[3]
    sheet.merge_cells('C17:F20')
    sheet['C21'] = reviews[4]
    sheet.merge_cells('C21:F24')

    sheet['G6'] = aboutReviews[0]
    sheet.merge_cells('G6:I6')
    sheet['G10'] = aboutReviews[1]
    sheet.merge_cells('G10:I10')
    sheet['G14'] = aboutReviews[2]
    sheet.merge_cells('G14:I14')
    sheet['G18'] = aboutReviews[3]
    sheet.merge_cells('G18:I18')
    sheet['G22'] = aboutReviews[4]
    sheet.merge_cells('G22:I22')

    sheet['G8'] = TFIDF[0]
    sheet.merge_cells('G8:I8')
    sheet['G12'] = TFIDF[1]
    sheet.merge_cells('G12:I12')
    sheet['G16'] = TFIDF[2]
    sheet.merge_cells('G16:I16')
    sheet['G20'] = TFIDF[3]
    sheet.merge_cells('G20:I20')
    sheet['G24'] = TFIDF[4]
    sheet.merge_cells('G24:I24')

    # for review in reviews:
    #     cell_reviews= 'C' + str(rows_reviews)
    #     sheet[cell_reviews] = review
    #     cell_reviews =+ increment

    # rows_token_lema = 7
    # for lematization in aboutReviews:
    #     cell_token_lema = 'G'+ str(rows_token_lema)
    #     sheet[cell_token_lema] = lematization
    #     cell_token_lema =+ increment

    # rows_vector = 9
    # for vector in TFIDF:
    #     cell_vector = 'G'+ str(rows_vector)
    #     sheet[cell_vector] = vector
    #     cell_vector =+ increment

    image = Image("graphs_result/"+placeID+".png")

    sheet.add_image(image, "L8")

    country = extraInformation[0]['short_country']
    PNG_country_path = Image("png_flags/"+country.lower() + '.png')
    PNG_country_path.width = 120
    PNG_country_path.height = 70
    sheet.add_image(PNG_country_path,'J3')

    sheet['H3'] = extraInformation[0]['Country']
    sheet.merge_cells("H3:I3")
    sheet['H2'] = extraInformation[0]['Website']
    sheet.merge_cells("H2:J2")
    sheet['B2'] = extraInformation[0]['Place_name']

    book.save("project.xlsx")
